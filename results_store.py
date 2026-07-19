"""
results_store.py
Handles saving candidate assessment results to an Excel workbook.
Results are split into Section A (JD-based) and Section B (Resume-based).
"""
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

RESULTS_DIR = "./results"
WORKBOOK_FILENAME = "IBM_Interview_Results.xlsx"

# ── Summary sheet headers ──────────────────────────────────────────────────────
SUMMARY_HEADERS = [
    "Submission #",
    "Submitted At (UTC)",
    "Candidate Name",
    "Candidate Email",
    "Candidate DOB",
    "Candidate Experience (Yrs)",
    "Interviewer Name",
    "Interviewer Email",
    "Difficulty",
    "Total Questions",
    # Section A
    "Sec A Questions",
    "Sec A Correct",
    "Sec A Score (%)",
    "Sec A Pass/Fail",
    # Section B
    "Sec B Questions",
    "Sec B Correct",
    "Sec B Score (%)",
    "Sec B Pass/Fail",
    "Status",
    "Webcam Video",
]

# ── Detail sheet headers ───────────────────────────────────────────────────────
DETAIL_HEADERS = [
    "Submission #",
    "Candidate Name",
    "Section",
    "Q#",
    "Question",
    "Source",
    "Candidate Answer",
    "Correct Answer",
    "Result",
]

# IBM Blue / Red header colours
_HDR_OVERALL = "1F3864"   # dark navy  — overall columns
_HDR_SEC_A   = "0F4E9C"   # IBM Blue   — Section A columns
_HDR_SEC_B   = "7B3F00"   # dark brown — Section B columns (distinct from A)
_HDR_DETAIL  = "1F4E79"   # slate blue — detail sheet


def get_results_workbook_path() -> str:
    os.makedirs(RESULTS_DIR, exist_ok=True)
    return os.path.join(RESULTS_DIR, WORKBOOK_FILENAME)


def check_duplicate_submission(candidate_name: str, candidate_email: str) -> str | None:
    """
    Check if a submission already exists for the given candidate name and email in the Excel results.
    Returns the formatted date string (e.g. '15 Jul 2026') if found, otherwise None.
    """
    if not _OPENPYXL_AVAILABLE:
        return None
    wb_path = get_results_workbook_path()
    if not os.path.exists(wb_path):
        return None
    try:
        wb = openpyxl.load_workbook(wb_path, data_only=True)
        if "Summary" not in wb.sheetnames:
            return None
        ws = wb["Summary"]
        target_name = candidate_name.strip().lower()
        target_email = candidate_email.strip().lower()
        
        # Iterate over rows starting from row 2
        for r in range(2, ws.max_row + 1):
            name_val = ws.cell(row=r, column=3).value
            email_val = ws.cell(row=r, column=4).value
            if name_val and email_val:
                if str(name_val).strip().lower() == target_name and str(email_val).strip().lower() == target_email:
                    date_val = ws.cell(row=r, column=2).value
                    if date_val:
                        try:
                            from datetime import datetime
                            # Handle datetime objects parsed by openpyxl
                            if isinstance(date_val, datetime):
                                return date_val.strftime("%d %b %Y")
                            # Parse ISO string
                            dt = datetime.fromisoformat(str(date_val))
                            return dt.strftime("%d %b %Y")
                        except Exception:
                            # Custom string parsing fallback (e.g. YYYY-MM-DD)
                            s = str(date_val)
                            if len(s) >= 10:
                                try:
                                    parts = s[:10].split("-")
                                    if len(parts) == 3:
                                        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
                                        year, month, day = parts[0], int(parts[1]), parts[2]
                                        return f"{int(day)} {months[month-1]} {year}"
                                except Exception:
                                    pass
                                return s[:10]
                            return s
                    return "an earlier date"
        return None
    except Exception as e:
        print(f"[WARN] Error checking duplicate submission: {e}")
        return None


def _create_fresh_workbook():
    """Create and return a new workbook with the two sheets pre-formatted."""
    if not _OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_header(ws_summary)

    # ── Detail sheet ──────────────────────────────────────────────────────────
    ws_detail = wb.create_sheet(title="Detailed Answers")
    _write_header_row(ws_detail, DETAIL_HEADERS, header_color=_HDR_DETAIL)

    return wb


def _write_header_row(ws, headers: list, header_color: str = _HDR_OVERALL):
    """Write a uniformly-styled header row."""
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor=header_color)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin
    ws.row_dimensions[1].height = 32


def _write_summary_header(ws):
    """Write Summary header with colour-coded section columns."""
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)

    colour_map = {
        # col index (1-based): fill colour
        1:  _HDR_OVERALL, 2:  _HDR_OVERALL, 3:  _HDR_OVERALL,
        4:  _HDR_OVERALL, 5:  _HDR_OVERALL, 6:  _HDR_OVERALL,
        7:  _HDR_OVERALL, 8:  _HDR_OVERALL, 9:  _HDR_OVERALL, 10: _HDR_OVERALL,
        11: _HDR_SEC_A,   12: _HDR_SEC_A,   13: _HDR_SEC_A,   14: _HDR_SEC_A,
        15: _HDR_SEC_B,   16: _HDR_SEC_B,   17: _HDR_SEC_B,   18: _HDR_SEC_B,
        19: _HDR_OVERALL, 20: _HDR_OVERALL,
    }
    for col_idx, header in enumerate(SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=colour_map.get(col_idx, _HDR_OVERALL))
        cell.alignment = align_c
        cell.border    = thin
    ws.row_dimensions[1].height = 32


def _auto_fit_columns(ws, min_width: int = 12, max_width: int = 60):
    """Roughly auto-fit column widths based on content."""
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len  = max(max_len, cell_len)
            except Exception:
                pass
        adjusted = max(min_width, min(max_width, max_len + 4))
        ws.column_dimensions[col_letter].width = adjusted


def _style_passfail(cell, passed: bool):
    if passed:
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
        cell.font = Font(bold=True, color="276221")
    else:
        cell.fill = PatternFill("solid", fgColor="FFC7CE")
        cell.font = Font(bold=True, color="9C0006")
    cell.alignment = Alignment(horizontal="center")


def save_candidate_result(
    interviewer_name: str,
    interviewer_email: str,
    candidate_name: str,
    candidate_email: str,
    difficulty: str,
    question_count: int,
    correct_count: int,
    score_percent: float,
    passed: bool,
    detailed_answers: list[dict],
    submitted_at: str,
    # Section-level data (with defaults for backward compat)
    count_a: int = 0,
    count_b: int = 0,
    correct_a: int = 0,
    correct_b: int = 0,
    score_a_percent: float = 0.0,
    score_b_percent: float = 0.0,
    candidate_dob: str = "",
    candidate_experience: str = "",
    status: str = "Completed",
    video_status: str = "N/A",
):
    """
    Append a candidate's result to the Excel workbook.
    Creates the workbook if it doesn't exist yet.
    """
    if not _OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb_path = get_results_workbook_path()

    if os.path.exists(wb_path):
        wb        = openpyxl.load_workbook(wb_path)
        ws_summary = wb["Summary"]
        ws_detail  = wb["Detailed Answers"]
    else:
        wb        = _create_fresh_workbook()
        ws_summary = wb["Summary"]
        ws_detail  = wb["Detailed Answers"]

    # ── Submission number ─────────────────────────────────────────────────────
    submission_num = ws_summary.max_row   # header = row 1, so this is next row index

    # ── Summary row ───────────────────────────────────────────────────────────
    pass_a_fail_str = "PASS" if score_a_percent >= 60 else "FAIL"
    pass_b_fail_str = "PASS" if score_b_percent >= 60 else "FAIL"
    if count_b == 0:
        pass_b_fail_str = "N/A"
        score_b_val = "N/A"
    else:
        score_b_val = score_b_percent

    summary_row = [
        submission_num,
        submitted_at,
        candidate_name,
        candidate_email,
        candidate_dob,
        candidate_experience,
        interviewer_name,
        interviewer_email,
        difficulty.capitalize(),
        question_count,
        # Section A
        count_a,
        correct_a,
        score_a_percent,
        pass_a_fail_str,
        # Section B
        count_b,
        correct_b,
        score_b_val,
        pass_b_fail_str,
        status,
        video_status,
    ]
    ws_summary.append(summary_row)

    last_row = ws_summary.max_row

    # Style Section A Pass/Fail (col 14)
    _style_passfail(ws_summary.cell(row=last_row, column=14), score_a_percent >= 60)

    # Style Section B Pass/Fail (col 18)
    if count_b > 0:
        _style_passfail(ws_summary.cell(row=last_row, column=18), score_b_percent >= 60)
    else:
        cell_b = ws_summary.cell(row=last_row, column=18)
        cell_b.alignment = Alignment(horizontal="center")

    # Centre score/count columns
    for col in (10, 11, 12, 13, 15, 16, 17, 19, 20):
        ws_summary.cell(row=last_row, column=col).alignment = Alignment(horizontal="center")

    # Section A score tint (col 13)
    sec_a_cell = ws_summary.cell(row=last_row, column=13)
    sec_a_cell.fill      = PatternFill("solid", fgColor="DDEEFF")
    sec_a_cell.font      = Font(bold=True, color="0F4E9C")
    sec_a_cell.alignment = Alignment(horizontal="center")

    # Section B score tint (col 17)
    sec_b_cell = ws_summary.cell(row=last_row, column=17)
    sec_b_cell.alignment = Alignment(horizontal="center")
    if count_b > 0:
        sec_b_cell.fill      = PatternFill("solid", fgColor="FFF0DC")
        sec_b_cell.font      = Font(bold=True, color="7B3F00")

    # Alternate row shading (even rows get a light tint)
    if submission_num % 2 == 0:
        row_fill = PatternFill("solid", fgColor="F0F4FF")
        for col_idx in range(1, len(SUMMARY_HEADERS) + 1):
            c = ws_summary.cell(row=last_row, column=col_idx)
            if c.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                c.fill = row_fill

    # ── Detail rows ───────────────────────────────────────────────────────────
    sec_a_answers = [a for a in detailed_answers if a.get("section") == "A"]
    sec_b_answers = [a for a in detailed_answers if a.get("section") != "A"]

    _SECTION_FILLS = {
        "A": ("DDEEFF", "0F4E9C"),   # blue tint for section A header rows
        "B": ("FFF0DC", "7B3F00"),   # amber tint for section B header rows
    }

    def _write_section_block(answers: list[dict], section_label: str, q_offset: int):
        divider_label = (
            f"--- Section {section_label}: {'Job Description Questions' if section_label == 'A' else 'Resume Questions'} ---"
        )
        fill_hex, font_hex = _SECTION_FILLS[section_label]
        ws_detail.append([submission_num, candidate_name, divider_label, "", "", "", "", "", ""])
        div_row = ws_detail.max_row
        for col_idx in range(1, 10):
            c = ws_detail.cell(row=div_row, column=col_idx)
            c.fill      = PatternFill("solid", fgColor=fill_hex)
            c.font      = Font(bold=True, color=font_hex, italic=True)
            c.alignment = Alignment(horizontal="left")

        for rel_idx, answer in enumerate(answers, start=1):
            result_str = "Correct" if answer["is_correct"] else "Incorrect"
            ws_detail.append([
                submission_num,
                candidate_name,
                f"Section {section_label}",
                q_offset + rel_idx,
                answer.get("question", ""),
                answer.get("source_doc", ""),
                answer.get("selected_option", ""),
                answer.get("correct_option", ""),
                result_str,
            ])
            det_row = ws_detail.max_row
            res_cell = ws_detail.cell(row=det_row, column=9)
            if answer["is_correct"]:
                res_cell.fill = PatternFill("solid", fgColor="C6EFCE")
                res_cell.font = Font(bold=True, color="276221")
            else:
                res_cell.fill = PatternFill("solid", fgColor="FFC7CE")
                res_cell.font = Font(bold=True, color="9C0006")
            res_cell.alignment = Alignment(horizontal="center")
            
            # Section label cell tint
            sec_cell = ws_detail.cell(row=det_row, column=3)
            sec_cell.fill = PatternFill("solid", fgColor=fill_hex)
            sec_cell.font = Font(bold=True, color=font_hex)
            sec_cell.alignment = Alignment(horizontal="center")

    _write_section_block(sec_a_answers, "A", 0)
    _write_section_block(sec_b_answers, "B", len(sec_a_answers))

    # ── Auto-fit and save ─────────────────────────────────────────────────────
    _auto_fit_columns(ws_summary)
    _auto_fit_columns(ws_detail)

    wb.save(wb_path)
    return wb_path
